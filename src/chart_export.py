"""Build all dashboard charts as native Excel charts in a multi-sheet workbook."""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Theme colours (ARGB hex without #) ────────────────────────────────────
_PURPLE = '7C3AED'
_CYAN   = '06B6D4'
_RED    = 'EF4444'
_GREEN  = '10B981'
_AMBER  = 'F59E0B'
_MAGENTA = 'D946EF'
_BLUE   = '3B82F6'
_ORANGE = 'F97316'
_PALETTE = [_PURPLE, _CYAN, _RED, _GREEN, _AMBER, _MAGENTA, _BLUE, _ORANGE]

_HEADER_FILL = PatternFill('solid', fgColor='1E1F35')
_HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
_TITLE_FONT  = Font(bold=True, size=13, color=_PURPLE)
_THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
)


def _write_table(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    """Write a DataFrame as a formatted table starting at start_row. Returns next free row."""
    # Headers
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = _THIN_BORDER
    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='center')
    # Auto-width
    for c_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(c_idx)
        max_len = max(len(str(df.columns[c_idx - 1])),
                      df.iloc[:, c_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)
    return start_row + len(df) + 1


def _add_title(ws, title: str, row: int = 1):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _TITLE_FONT


def _style_chart(chart, title: str, width: int = 18, height: int = 12):
    chart.title = title
    chart.width = width
    chart.height = height
    chart.legend.position = 'b'
    chart.style = 10


def _pie(ws, df: pd.DataFrame, title: str, cat_col: int, val_col: int,
         data_start: int, data_end: int, anchor: str):
    """Add a pie chart to the worksheet."""
    chart = PieChart()
    chart.title = title
    chart.width = 16
    chart.height = 12
    chart.style = 10

    data = Reference(ws, min_col=val_col, min_row=data_start, max_row=data_end)
    cats = Reference(ws, min_col=cat_col, min_row=data_start + 1, max_row=data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Colour slices
    for i in range(data_end - data_start):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = _PALETTE[i % len(_PALETTE)]
        chart.series[0].data_points.append(pt)

    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showPercent = True
    chart.series[0].dLbls.showCatName = True

    ws.add_chart(chart, anchor)


def _bar(ws, df: pd.DataFrame, title: str, cat_col: int,
         val_cols: list[int], data_start: int, data_end: int,
         anchor: str, colours: list[str] | None = None, width: int = 18):
    """Add a bar chart."""
    chart = BarChart()
    _style_chart(chart, title, width=width)
    chart.type = 'col'
    chart.grouping = 'clustered'

    cats = Reference(ws, min_col=cat_col, min_row=data_start + 1, max_row=data_end)
    chart.set_categories(cats)

    for i, vc in enumerate(val_cols):
        data = Reference(ws, min_col=vc, min_row=data_start, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        if colours and i < len(colours):
            chart.series[i].graphicalProperties.solidFill = colours[i]

    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showVal = True

    ws.add_chart(chart, anchor)


def _hbar(ws, df: pd.DataFrame, title: str, cat_col: int,
          val_cols: list[int], data_start: int, data_end: int,
          anchor: str, colours: list[str] | None = None, width: int = 18, height: int = 12):
    """Add a horizontal bar chart."""
    chart = BarChart()
    _style_chart(chart, title, width=width, height=height)
    chart.type = 'bar'
    chart.grouping = 'clustered'

    cats = Reference(ws, min_col=cat_col, min_row=data_start + 1, max_row=data_end)
    chart.set_categories(cats)

    for i, vc in enumerate(val_cols):
        data = Reference(ws, min_col=vc, min_row=data_start, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        if colours and i < len(colours):
            chart.series[i].graphicalProperties.solidFill = colours[i]

    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showVal = True

    ws.add_chart(chart, anchor)


def _line(ws, title: str, cat_col: int, val_cols: list[int],
          data_start: int, data_end: int, anchor: str,
          colours: list[str] | None = None, width: int = 20, height: int = 12):
    """Add a line chart."""
    chart = LineChart()
    _style_chart(chart, title, width=width, height=height)

    cats = Reference(ws, min_col=cat_col, min_row=data_start + 1, max_row=data_end)
    chart.set_categories(cats)

    for i, vc in enumerate(val_cols):
        data = Reference(ws, min_col=vc, min_row=data_start, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        if colours and i < len(colours):
            chart.series[i].graphicalProperties.line.solidFill = colours[i]

    ws.add_chart(chart, anchor)


def build_charts_excel(filtered_df: pd.DataFrame, kpis: dict) -> io.BytesIO:
    """
    Generate an Excel workbook with data tables + native Excel charts per section.
    Returns a BytesIO buffer ready for st.download_button.
    """
    departed_df = filtered_df[filtered_df['Employee Status'] == 'Departed']
    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. KPI Summary ────────────────────────────────────────────────────
    ws = wb.create_sheet('KPI Summary')
    kpi_data = pd.DataFrame([
        {'Metric': 'Total Employees', 'Value': kpis['total']},
        {'Metric': 'Active Employees', 'Value': kpis['active']},
        {'Metric': 'Departed Employees', 'Value': kpis['departed']},
        {'Metric': 'Departure Rate %', 'Value': round(kpis['attrition_rate'], 1)},
        {'Metric': 'Retention Rate %', 'Value': round(kpis['retention_rate'], 1)},
        {'Metric': 'Avg Tenure (Months)', 'Value': round(kpis['avg_tenure'], 1)},
        {'Metric': 'Avg Age', 'Value': round(kpis['avg_age'], 0) if not pd.isna(kpis['avg_age']) else 'N/A'},
        {'Metric': 'Gender Ratio (M:F)', 'Value': kpis['gender_ratio']},
    ])
    _add_title(ws, 'Key Performance Indicators')
    _write_table(ws, kpi_data, start_row=3)

    # ── 2. Gender Distribution ────────────────────────────────────────────
    ws = wb.create_sheet('Gender Distribution')
    gender_df = filtered_df['Gender'].value_counts().reset_index()
    gender_df.columns = ['Gender', 'Count']
    _add_title(ws, 'Gender Distribution')
    end = _write_table(ws, gender_df, start_row=3)
    _pie(ws, gender_df, 'Gender Distribution', cat_col=1, val_col=2,
         data_start=3, data_end=end - 1, anchor='D3')

    # ── 3. Employment Status ──────────────────────────────────────────────
    ws = wb.create_sheet('Employment Status')
    status_df = filtered_df['Employee Status'].value_counts().reset_index()
    status_df.columns = ['Status', 'Count']
    _add_title(ws, 'Employment Status')
    end = _write_table(ws, status_df, start_row=3)
    _pie(ws, status_df, 'Employment Status', cat_col=1, val_col=2,
         data_start=3, data_end=end - 1, anchor='D3')

    # ── 4. Department Breakdown ───────────────────────────────────────────
    ws = wb.create_sheet('Department Breakdown')
    dept_status = filtered_df.groupby('Department')['Employee Status'].value_counts().unstack(fill_value=0).reset_index()
    _add_title(ws, 'Department Breakdown (Active vs Departed)')
    end = _write_table(ws, dept_status, start_row=3)
    val_cols = [i + 2 for i in range(len(dept_status.columns) - 1)]
    colours = [_CYAN, _RED]
    _bar(ws, dept_status, 'Department Breakdown', cat_col=1, val_cols=val_cols,
         data_start=3, data_end=end - 1, anchor=f'A{end + 1}', colours=colours, width=22)

    # ── 5. Exit Types ─────────────────────────────────────────────────────
    if len(departed_df) > 0:
        ws = wb.create_sheet('Exit Types')
        exit_df = departed_df['Exit Type'].value_counts().reset_index()
        exit_df.columns = ['Exit Type', 'Count']
        _add_title(ws, 'Exit Types')
        end = _write_table(ws, exit_df, start_row=3)
        _pie(ws, exit_df, 'Exit Types', cat_col=1, val_col=2,
             data_start=3, data_end=end - 1, anchor='D3')

    # ── 6. Exit Reason Categories ─────────────────────────────────────────
    if len(departed_df) > 0 and 'Exit Reason Category' in departed_df.columns:
        ws = wb.create_sheet('Exit Reasons')
        reason_df = departed_df['Exit Reason Category'].value_counts().reset_index()
        reason_df.columns = ['Category', 'Count']
        _add_title(ws, 'Exit Reason Categories')
        end = _write_table(ws, reason_df, start_row=3)
        _hbar(ws, reason_df, 'Exit Reason Categories', cat_col=1, val_cols=[2],
              data_start=3, data_end=end - 1, anchor=f'D3',
              colours=[_PURPLE], height=max(10, len(reason_df) * 1.5))

    # ── 7. Departure Rate by Department ──────────────────────────────────
    dept_attrition = filtered_df.groupby('Department').agg(
        Active=('Employee Status', lambda x: (x == 'Active').sum()),
        Departed=('Employee Status', lambda x: (x == 'Departed').sum()),
        Total=('Employee Status', 'count')
    ).reset_index()
    dept_attrition['Departure Rate %'] = (
        dept_attrition['Departed'] / dept_attrition['Total'] * 100
    ).round(1)
    dept_attrition = dept_attrition.sort_values('Departure Rate %', ascending=False)

    ws = wb.create_sheet('Dept Departure Rate')
    _add_title(ws, 'Departure Rate by Department')
    end = _write_table(ws, dept_attrition, start_row=3)
    _bar(ws, dept_attrition, 'Departure Rate %', cat_col=1, val_cols=[5],
         data_start=3, data_end=end - 1, anchor=f'A{end + 1}', colours=[_RED], width=22)

    # ── 8. Tenure by Department ───────────────────────────────────────────
    if 'Tenure (Months)' in filtered_df.columns:
        tenure_dept = (
            filtered_df.groupby('Department')['Tenure (Months)']
            .agg(['mean', 'median', 'count']).round(1)
            .rename(columns={'mean': 'Avg Tenure', 'median': 'Median Tenure', 'count': 'Count'})
            .reset_index().sort_values('Avg Tenure', ascending=False)
        )
        ws = wb.create_sheet('Tenure by Dept')
        _add_title(ws, 'Average Tenure by Department')
        end = _write_table(ws, tenure_dept, start_row=3)
        _bar(ws, tenure_dept, 'Avg Tenure (Months)', cat_col=1, val_cols=[2, 3],
             data_start=3, data_end=end - 1, anchor=f'A{end + 1}',
             colours=[_CYAN, _AMBER], width=22)

    # ── 9. Time-to-Departure ──────────────────────────────────────────────
    if len(departed_df) > 0 and 'Tenure (Months)' in departed_df.columns:
        ttd = (
            departed_df.groupby('Department')['Tenure (Months)']
            .agg(Avg='mean', Median='median', Count='count')
            .round(1).reset_index().sort_values('Avg')
        )
        ws = wb.create_sheet('Time to Departure')
        _add_title(ws, 'Average Tenure at Exit by Department')
        end = _write_table(ws, ttd, start_row=3)
        _hbar(ws, ttd, 'Avg Months Before Departure', cat_col=1, val_cols=[2],
              data_start=3, data_end=end - 1, anchor='F3',
              colours=[_GREEN], height=max(10, len(ttd) * 1.5))

    # ── 10. Early Departure Rate ──────────────────────────────────────────
    if len(departed_df) > 0 and 'Tenure (Months)' in departed_df.columns:
        dept_stats = departed_df.groupby('Department').agg(
            Total_Departed=('Tenure (Months)', 'count'),
            Early_Departed=('Tenure (Months)', lambda x: (x <= 3).sum()),
        ).reset_index()
        dept_stats['Early Departure Rate %'] = (
            dept_stats['Early_Departed'] / dept_stats['Total_Departed'] * 100
        ).round(1)
        dept_stats = dept_stats.sort_values('Early Departure Rate %', ascending=False)

        ws = wb.create_sheet('Early Departure Rate')
        _add_title(ws, 'Early Departure Rate (Left ≤3 Months)')
        end = _write_table(ws, dept_stats, start_row=3)
        _bar(ws, dept_stats, 'Early Departure Rate %', cat_col=1, val_cols=[4],
             data_start=3, data_end=end - 1, anchor=f'A{end + 1}',
             colours=[_AMBER], width=22)

    # ── 11. Monthly Hiring vs Departures ─────────────────────────────────
    if 'Join Month' in filtered_df.columns:
        hiring = filtered_df.groupby('Join Month').size().rename('Hires')
        exits = (
            departed_df.groupby('Exit Month').size().rename('Exits')
            if len(departed_df) > 0 and 'Exit Month' in departed_df.columns
            else pd.Series(dtype=int)
        )
        all_months = sorted(set(hiring.index.tolist() + exits.index.tolist()))
        combined = pd.DataFrame({'Month': all_months})
        combined['Hires'] = combined['Month'].map(hiring).fillna(0).astype(int)
        combined['Exits'] = combined['Month'].map(exits).fillna(0).astype(int)
        combined['Net'] = combined['Hires'] - combined['Exits']

        ws = wb.create_sheet('Hiring vs Departures')
        _add_title(ws, 'Monthly Hiring vs Departures')
        end = _write_table(ws, combined, start_row=3)
        _line(ws, 'Hiring vs Departures', cat_col=1, val_cols=[2, 3],
              data_start=3, data_end=end - 1, anchor=f'A{end + 1}',
              colours=[_GREEN, _RED], width=24, height=14)

    # ── 12. Headcount by Year ─────────────────────────────────────────────
    if 'Join Year' in filtered_df.columns:
        headcount = (
            filtered_df[filtered_df['Join Year'] > 2000]
            .groupby('Join Year')['Employee Status'].value_counts()
            .unstack(fill_value=0).reset_index()
        )
        if len(headcount) > 0:
            ws = wb.create_sheet('Headcount by Year')
            _add_title(ws, 'Headcount Summary by Year')
            end = _write_table(ws, headcount, start_row=3)
            val_cols = [i + 2 for i in range(len(headcount.columns) - 1)]
            chart = BarChart()
            _style_chart(chart, 'Headcount by Year', width=22)
            chart.grouping = 'stacked'
            cats = Reference(ws, min_col=1, min_row=4, max_row=end - 1)
            chart.set_categories(cats)
            for i, vc in enumerate(val_cols):
                data = Reference(ws, min_col=vc, min_row=3, max_row=end - 1)
                chart.add_data(data, titles_from_data=True)
                chart.series[i].graphicalProperties.solidFill = [_GREEN, _RED][i % 2]
            ws.add_chart(chart, f'A{end + 1}')

    # ── 13. Vendor Analysis ───────────────────────────────────────────────
    if 'Vendor' in filtered_df.columns:
        vendor_df = filtered_df['Vendor'].value_counts().reset_index()
        vendor_df.columns = ['Vendor', 'Count']
        ws = wb.create_sheet('Vendor Analysis')
        _add_title(ws, 'Vendor / Source Analysis')
        end = _write_table(ws, vendor_df, start_row=3)
        _pie(ws, vendor_df, 'Vendor Distribution', cat_col=1, val_col=2,
             data_start=3, data_end=end - 1, anchor='D3')

    # ── Save ──────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
